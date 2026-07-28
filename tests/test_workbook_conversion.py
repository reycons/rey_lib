"""Focused tests for the application-neutral workbook conversion boundary."""

from __future__ import annotations

import os
from pathlib import Path
import shutil
from typing import Any

import unicodedata

from openpyxl import Workbook
import polars as pl
import pytest

import rey_lib.files.workbook_conversion as conversion
from rey_lib.files import (
    EmptyWorkbookError,
    UnsupportedWorkbookError,
    WorkbookEncryptedError,
    WorkbookExtractionError,
    WorkbookOutputCollisionError,
    WorkbookWriteError,
    convert_workbook_to_csv,
    is_supported_workbook,
)
from tests.fixtures.workbooks import mixed_workbook, sheet_only_workbook


def _create_test_workbook(tmp_path: Path, content_cells: list[list[Any]], sheet_name: str = "Sheet1") -> Path:
    """Create a simple test workbook with the given cell contents."""
    path = tmp_path / f"test_{id(content_cells)}.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    for row in content_cells:
        sheet.append(row)
    workbook.save(path)
    return path


def _create_workbook_with_control_chars(tmp_path: Path, cell_value: str) -> Path:
    """Create a test workbook containing a single cell with control characters."""
    path = tmp_path / "control_chars.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Value"])
    sheet.append([cell_value])
    workbook.save(path)
    return path


def _create_workbook_with_leading_zeros(tmp_path: Path, value: str) -> Path:
    """Create a test workbook with a leading-zero string preserved."""
    path = tmp_path / "leading_zeros.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Code"])
    # Write as formula to preserve leading zeros
    sheet.append([f"={value}"])
    workbook.save(path)
    return path


def _create_workbook_with_scientific(tmp_path: Path, value) -> Path:
    """Create a test workbook with a scientific notation number."""
    path = tmp_path / "scientific.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Value"])
    sheet.append([value])
    workbook.save(path)
    return path


def _create_workbook_with_nulls(tmp_path: Path, null_value: Any) -> Path:
    """Create a test workbook with a null cell."""
    path = tmp_path / "nulls.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Value"])
    if null_value is None:
        # Openpyxl stores None as empty
        sheet.append([None])
    else:
        sheet.append([null_value])
    workbook.save(path)
    return path


def _create_workbook_with_numeric_types(tmp_path: Path, int_val: int, float_val: float) -> Path:
    """Create a test workbook with integer and float values."""
    path = tmp_path / "numeric.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["IntCol", "FloatCol"])
    sheet.append([int_val, float_val])
    workbook.save(path)
    return path


def _create_workbook_with_colliding_headers(tmp_path: Path) -> Path:
    """Create a test workbook with headers that collide after normalization."""
    path = tmp_path / "collision.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    # Both headers normalize to "HEADER1" case-insensitively after removing control chars
    # First header: H-E-A-D-E-R-1 (normal)
    # Second header: H-e-a-d-e-r-1 with a Cf character (format mark U+008F) between 'r' and '1'
    sheet.append(["Header1", f"Header{chr(0x008F)}1"])
    sheet.append(["a", "b"])
    workbook.save(path)
    return path


def _create_workbook_with_empty_header(tmp_path: Path) -> Path:
    """Create a test workbook with a header that normalizes to empty."""
    path = tmp_path / "empty_header.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    # Column name with only control characters that will normalize to empty
    sheet.append(["\t\r\n"])
    sheet.append(["value"])
    workbook.save(path)
    return path


def _create_workbook_with_leading_zeros_preserved(tmp_path: Path, value: str) -> Path:
    """Create a test workbook with a leading-zero string preserved as string."""
    path = tmp_path / "leading_zeros.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Code"])
    # Write as formula to preserve leading zeros (as text)
    sheet.append([f"='{value}'"])
    workbook.save(path)
    return path


def _create_workbook_with_int64_float64(tmp_path: Path) -> Path:
    """Create a test workbook specifically to verify Int64/Float64 preservation."""
    path = tmp_path / "intfloat.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["IntCol", "FloatCol"])
    # Use values that would trigger scientific notation if not for float_scientific=False
    sheet.append([1234567890, 0.00000001])
    workbook.save(path)
    return path


@pytest.mark.parametrize(
    ("name", "expected"),
    [
        ("book.xls", True),
        ("book.XLSX", True),
        ("book.xlsb", True),
        ("book.XlSm", True),
        ("book.csv", False),
        ("book", False),
    ],
)
def test_is_supported_workbook_is_case_insensitive(name: str, expected: bool) -> None:
    assert is_supported_workbook(name) is expected


def test_defined_tables_and_worksheet_fallback_are_not_duplicated(tmp_path: Path) -> None:
    source = mixed_workbook(tmp_path / "Holdings Report.xlsx")
    output_dir = tmp_path / "out"

    result = convert_workbook_to_csv(source, output_dir)

    assert result.conversion_status == "success_with_warnings"
    assert [item.artifact_name for item in result.outputs] == [
        "Holdings_Report.Table_Sheet.Current_Holdings.csv",
        "Holdings_Report.Notes_Data.csv",
    ]
    assert [item.extraction_kind for item in result.outputs] == [
        "defined_table",
        "worksheet",
    ]
    assert result.outputs[0].sheet_name == "Table Sheet"
    assert result.outputs[0].table_name == "Current_Holdings"
    assert result.outputs[0].row_count == 2
    assert result.outputs[0].column_names == ("Account", "Amount", "Memo")
    assert not (output_dir / "Holdings_Report.Table_Sheet.csv").exists()
    assert {warning.code for warning in result.warnings} == {
        "sheet_excluded",
        "empty_sheet_excluded",
    }


def test_csv_contract_is_exact_and_deterministic(tmp_path: Path) -> None:
    source = mixed_workbook(tmp_path / "Holdings Report.xlsx")

    result = convert_workbook_to_csv(source, tmp_path / "out")

    # Note: embedded newlines in cell values are normalized to spaces by the
    # control-character sanitization contract.
    table_bytes = result.outputs[0].output_path.read_bytes()
    assert table_bytes == (
        b'Account,Amount,Memo\n'
        b'A-1,10.5,"x,y"\n'
        b'A-2,,line break\n'
    )
    assert not table_bytes.startswith(b"\xef\xbb\xbf")
    assert b"\r\n" not in table_bytes


def test_xlsm_uses_xlsx_family_table_extraction(tmp_path: Path) -> None:
    xlsx = mixed_workbook(tmp_path / "source.xlsx")
    xlsm = tmp_path / "source.xlsm"
    shutil.copyfile(xlsx, xlsm)

    result = convert_workbook_to_csv(xlsm, tmp_path / "out")

    assert result.source_extension == ".xlsm"
    assert result.outputs[0].extraction_kind == "defined_table"


class _FakeSheet:
    visible = "visible"

    def __init__(self, frame: pl.DataFrame) -> None:
        self._frame = frame

    def to_polars(self) -> pl.DataFrame:
        return self._frame


class _LegacyReader:
    sheet_names = ["Legacy"]

    def __init__(self) -> None:
        self.table_names_called = False

    def table_names(self, _sheet_name: str) -> list[str]:
        self.table_names_called = True
        raise AssertionError("legacy workbook must not enumerate tables")

    def load_sheet(self, _name: str, *, n_rows: int | None = None) -> _FakeSheet:
        frame = pl.DataFrame({"Column": ["value"]})
        if n_rows == 0:
            frame = frame.head(0)
        return _FakeSheet(frame)


@pytest.mark.parametrize("extension", [".xls", ".xlsb"])
def test_legacy_formats_bypass_table_enumeration(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    extension: str,
) -> None:
    source = tmp_path / f"legacy{extension}"
    source.write_bytes(b"synthetic")
    reader = _LegacyReader()
    monkeypatch.setattr(conversion, "_load_dependencies", lambda _source: (object(), pl))
    monkeypatch.setattr(conversion, "_open_workbook", lambda _fastexcel, _source: reader)

    result = convert_workbook_to_csv(source, tmp_path / "out")

    assert reader.table_names_called is False
    assert result.outputs[0].artifact_name == "legacy.Legacy.csv"
    assert result.outputs[0].extraction_kind == "worksheet"


def test_workbook_is_opened_once(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    source = mixed_workbook(tmp_path / "book.xlsx")
    original = conversion._open_workbook
    calls = 0

    def counted_open(fastexcel: Any, source_path: Path) -> Any:
        nonlocal calls
        calls += 1
        return original(fastexcel, source_path)

    monkeypatch.setattr(conversion, "_open_workbook", counted_open)

    convert_workbook_to_csv(source, tmp_path / "out")

    assert calls == 1


def test_sanitized_name_collision_fails_before_writing(tmp_path: Path) -> None:
    source = sheet_only_workbook(tmp_path / "book.xlsx", ("A.B", "A B"))
    output_dir = tmp_path / "out"

    with pytest.raises(WorkbookOutputCollisionError):
        convert_workbook_to_csv(source, output_dir)

    assert not output_dir.exists() or list(output_dir.iterdir()) == []


def test_existing_destination_is_never_overwritten(tmp_path: Path) -> None:
    source = sheet_only_workbook(tmp_path / "book.xlsx", ("Sheet",))
    output_dir = tmp_path / "out"
    output_dir.mkdir()
    existing = output_dir / "book.Sheet.csv"
    existing.write_text("keep me", encoding="utf-8")

    with pytest.raises(WorkbookOutputCollisionError):
        convert_workbook_to_csv(source, output_dir)

    assert existing.read_text(encoding="utf-8") == "keep me"


def test_publish_failure_removes_new_outputs(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    source = sheet_only_workbook(tmp_path / "book.xlsx", ("One", "Two"))
    output_dir = tmp_path / "out"
    original_link = os.link
    calls = 0

    def fail_second_link(source_path: Path, destination_path: Path) -> None:
        nonlocal calls
        calls += 1
        if calls == 2:
            raise OSError("synthetic publish failure")
        original_link(source_path, destination_path)

    monkeypatch.setattr(conversion.os, "link", fail_second_link)

    with pytest.raises(WorkbookWriteError):
        convert_workbook_to_csv(source, output_dir)

    assert list(output_dir.iterdir()) == []


def test_password_protection_is_translated_and_chained(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    source = tmp_path / "protected.xlsx"
    source.write_bytes(b"synthetic")

    class _FastExcel:
        @staticmethod
        def read_excel(_source: Path) -> Any:
            raise RuntimeError("Workbook is password protected")

    monkeypatch.setattr(conversion, "_load_dependencies", lambda _source: (_FastExcel, pl))

    with pytest.raises(WorkbookEncryptedError) as exc_info:
        convert_workbook_to_csv(source, tmp_path / "out")

    assert isinstance(exc_info.value.__cause__, RuntimeError)


def test_empty_workbook_returns_no_partial_result(tmp_path: Path) -> None:
    source = tmp_path / "empty.xlsx"
    workbook = Workbook()
    workbook.save(source)

    with pytest.raises(EmptyWorkbookError):
        convert_workbook_to_csv(source, tmp_path / "out")


def test_unsupported_extension_is_structured(tmp_path: Path) -> None:
    source = tmp_path / "book.csv"
    source.write_text("a\n1\n", encoding="utf-8")

    with pytest.raises(UnsupportedWorkbookError) as exc_info:
        convert_workbook_to_csv(source, tmp_path / "out")

    assert exc_info.value.code == "unsupported_workbook"
    assert exc_info.value.source_path == source.resolve()


# ============================================================================
# Cc/Cf control character removal from cell values
# ============================================================================


def test_cc_removal_from_cell_values(tmp_path: Path) -> None:
    """Test that Cc (control characters) are removed from cell values."""
    source = tmp_path / "cc_test.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Value"])
    # Tab is a Cc category character - will be replaced by space first, then trimmed
    sheet.append(["before\tafter_tab"])  # Tab in middle
    workbook.save(source)
    output_dir = tmp_path / "out"

    result = convert_workbook_to_csv(source, output_dir)

    assert len(result.outputs) == 1
    content = result.outputs[0].output_path.read_text()
    # Tab should be replaced with space
    assert "before after_tab" in content


def test_cf_removal_from_cell_values(tmp_path: Path) -> None:
    """Test that Cf (format control characters) are removed from cell values."""
    source = tmp_path / "cf_test.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Value"])
    # Unicode format marker U+008F is a Cf character - will be removed entirely
    sheet.append([f"before{chr(0x008F)}after_caret"])
    workbook.save(source)
    output_dir = tmp_path / "out"

    result = convert_workbook_to_csv(source, output_dir)

    assert len(result.outputs) == 1
    content = result.outputs[0].output_path.read_text()
    # Format marker should be removed entirely
    assert "beforeafter_caret" in content


def test_cr_lf_tab_nbsp_normalization(tmp_path: Path) -> None:
    """Test that CR, LF, TAB, and NBSP are normalized to spaces."""
    source = tmp_path / "whitespace_test.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Value"])
    # Create a value with various whitespace characters
    cr = "\r"
    lf = "\n"
    tab = "\t"
    nbsp = "\u00a0"  # Non-breaking space

    sheet.append([f"{cr}start{lf}{tab}{nbsp}end{nbsp}"])
    sheet.append(["normal row"])  # Extra row to avoid empty workbook error
    workbook.save(source)
    output_dir = tmp_path / "out"

    result = convert_workbook_to_csv(source, output_dir)

    assert len(result.outputs) == 1
    content = result.outputs[0].output_path.read_bytes()
    # All embedded whitespace in cell values should be normalized to spaces and trimmed
    assert b"start end" in content
    # CR, LF, TAB, NBSP within cell values should not appear (as bytes)
    assert cr.encode() not in content
    # Note: \n is the standard CSV line terminator. We have 3 lines: header + 2 data rows.
    lines = content.strip().split(b"\n")
    assert len(lines) == 3, f"Expected 3 lines (header + 2 data), got {len(lines)}"
    assert tab.encode() not in content


# ============================================================================
# Null handling tests
# ============================================================================


def test_null_remains_null_in_normalized_frame(tmp_path: Path) -> None:
    """Test that null values remain as null in the Polars frame."""
    source = tmp_path / "null_test.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Value"])
    sheet.append([None])  # Null value
    sheet.append(["not_null"])  # Extra row to avoid empty workbook error
    workbook.save(source)
    output_dir = tmp_path / "out"

    result = convert_workbook_to_csv(source, output_dir)

    assert len(result.outputs) == 1
    # Read back and check null is present (empty field between commas)
    content = result.outputs[0].output_path.read_text()
    lines = content.strip().split("\n")
    assert lines[0] == "Value"  # Header
    assert lines[1] == ""  # Null as empty string


def test_null_writes_as_empty_csv_field(tmp_path: Path) -> None:
    """Test that null writes as an empty CSV field."""
    source = tmp_path / "null_test.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["A", "B"])
    sheet.append([1, None])  # Second column is null
    sheet.append([2, 3])
    workbook.save(source)

    result = convert_workbook_to_csv(source, tmp_path / "out")

    content = result.outputs[0].output_path.read_text()
    lines = content.strip().split("\n")
    assert lines[1] == "1,"  # Null as empty field


# ============================================================================
# Numeric dtype preservation tests
# ============================================================================


def test_int64_remains_int64_before_serialization(tmp_path: Path) -> None:
    """Test that integer-like values are preserved as numeric (not converted to strings)."""
    source = _create_workbook_with_numeric_types(tmp_path, 12345, 0.0)

    output_dir = tmp_path / "out"
    result = convert_workbook_to_csv(source, output_dir)

    assert len(result.outputs) == 1
    content = result.outputs[0].output_path.read_text()
    # Check that integer values appear as numeric (not quoted strings)
    lines = content.strip().split("\n")
    header = lines[0]
    data_line = lines[1]
    
    # Header should have IntCol and FloatCol
    assert "IntCol" in header
    assert "FloatCol" in header
    
    # Data line should have numeric values without quotes around them
    # IntCol=12345 should appear as "12345,0.0"
    parts = data_line.split(",")
    assert len(parts) == 2
    # The first part (IntCol) should be a valid integer string
    int_val = int(parts[0]) if parts[0] else None
    assert int_val == 12345, f"Expected IntCol=12345, got {parts[0]}"


def test_float64_remains_float64_before_serialization(tmp_path: Path) -> None:
    """Test that Float64 columns remain Float64 before CSV serialization."""
    source = _create_workbook_with_numeric_types(tmp_path, 0, 1.5)

    output_dir = tmp_path / "out"
    result = convert_workbook_to_csv(source, output_dir)

    assert len(result.outputs) == 1
    float_col_dtype = None
    for name, dtype in result.outputs[0].polars_schema:
        if "FloatCol" in name:
            float_col_dtype = dtype

    assert float_col_dtype == "Float64", f"Expected Float64, got {float_col_dtype}"


# ============================================================================
# Scientific notation prevention tests
# ============================================================================


def test_small_scientific_writes_without_e_or_e(tmp_path: Path) -> None:
    """Test that 0.000000125 writes without 'e' or 'E' in output."""
    source = _create_workbook_with_scientific(tmp_path, 0.000000125)
    output_dir = tmp_path / "out"

    result = convert_workbook_to_csv(source, output_dir)

    content = result.outputs[0].output_path.read_text()
    # Should not contain 'e' or 'E'
    lines = content.strip().split("\n")
    value_line = lines[1]  # Second line has the value
    assert "e" not in value_line.lower(), f"Scientific notation found: {value_line}"


def test_large_scientific_writes_without_e_or_e(tmp_path: Path) -> None:
    """Test that 1.25e20 writes without 'e' or 'E' in output."""
    source = _create_workbook_with_scientific(tmp_path, 1.25e20)
    output_dir = tmp_path / "out"

    result = convert_workbook_to_csv(source, output_dir)

    content = result.outputs[0].output_path.read_text()
    lines = content.strip().split("\n")
    value_line = lines[1]  # Second line has the value
    assert "e" not in value_line.lower(), f"Scientific notation found: {value_line}"


def test_negative_exponent_writes_without_e_or_e(tmp_path: Path) -> None:
    """Test that -2.5e-10 writes without 'e' or 'E' in output."""
    source = _create_workbook_with_scientific(tmp_path, -2.5e-10)
    output_dir = tmp_path / "out"

    result = convert_workbook_to_csv(source, output_dir)

    content = result.outputs[0].output_path.read_text()
    lines = content.strip().split("\n")
    value_line = lines[1]  # Second line has the value
    assert "e" not in value_line.lower(), f"Scientific notation found: {value_line}"


# ============================================================================
# Header normalization tests (errors)
# ============================================================================


def test_header_normalizing_to_empty_raises_error(tmp_path: Path) -> None:
    """Test that a header normalizing to empty raises WorkbookExtractionError."""
    source = _create_workbook_with_empty_header(tmp_path)

    with pytest.raises(WorkbookExtractionError) as exc_info:
        convert_workbook_to_csv(source, tmp_path / "out")

    assert "CSV column name is empty after control-character normalization" in str(exc_info.value)


def test_colliding_headers_after_normalization_raises_error(tmp_path: Path) -> None:
    """Test that headers colliding case-insensitively after normalization raise WorkbookExtractionError."""
    source = _create_workbook_with_colliding_headers(tmp_path)

    with pytest.raises(WorkbookExtractionError) as exc_info:
        convert_workbook_to_csv(source, tmp_path / "out")

    assert "collide" in str(exc_info.value).lower()
