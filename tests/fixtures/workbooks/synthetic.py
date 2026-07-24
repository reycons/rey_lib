"""Create small synthetic XLSX-family workbooks for conversion tests."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.worksheet.table import Table


def mixed_workbook(path: Path) -> Path:
    """Create a workbook containing a table, fallback, hidden, and empty sheet."""

    workbook = Workbook()
    table_sheet = workbook.active
    table_sheet.title = "Table Sheet"
    table_sheet.append(["Account", "Amount", "Memo"])
    table_sheet.append(["A-1", 10.5, "x,y"])
    table_sheet.append(["A-2", None, "line\nbreak"])
    table_sheet.add_table(Table(displayName="Current_Holdings", ref="A1:C3"))

    fallback = workbook.create_sheet("Notes.Data")
    fallback.append(["Code", "Enabled"])
    fallback.append(["N1", True])

    hidden = workbook.create_sheet("Hidden Sheet")
    hidden.append(["Secret"])
    hidden.append(["not emitted"])
    hidden.sheet_state = "hidden"

    workbook.create_sheet("Empty Sheet")
    workbook.save(path)
    return path


def sheet_only_workbook(path: Path, sheet_names: tuple[str, ...]) -> Path:
    """Create a workbook containing simple fallback worksheets."""

    workbook = Workbook()
    first = workbook.active
    first.title = sheet_names[0]
    first.append(["Value"])
    first.append([1])
    for sheet_name in sheet_names[1:]:
        sheet = workbook.create_sheet(sheet_name)
        sheet.append(["Value"])
        sheet.append([2])
    workbook.save(path)
    return path
